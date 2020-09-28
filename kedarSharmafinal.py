import cv2
import openpyxl as op

# IMP_Note: when it will ask to point where you want to put the text on template,
# this pointer is pointing at the left-bottom of the text
# for example, i want to put 'Late SSR', then pointer is pointing at left-bottom of letter 'L'

# below code is deciding where we want to put the date that we have in each row
img = cv2.imread('template5.jpeg')  # loading the certificate template
arr = []


def click_event(event, x, y, flags, param):  # flags: Any relevant flags passed by OpenCV.
    # params: Any extra parameters supplied by OpenCV
    if event == cv2.EVENT_LBUTTONDOWN:
       # print(x, ', ', y)
        arr.append((x, y))
        #font = cv2.FONT_HERSHEY_SCRIPT_COMPLEX
        #strXY = str(x) + ', ' + str(y)
        #cv2.putText(img, strXY, (x, y), font, .4, (0, 0, 0), 1)
        cv2.imshow('image', img)


cv2.imshow('image', img)
cv2.setMouseCallback('image', click_event)

cv2.waitKey(0)
#print(arr)
cv2.destroyAllWindows()

# below code is putting the details on template and saving the certificates at the destination path

dest_path = 'generated'  # giving the destination folder the certificates are to be stored
sheet = op.load_workbook('data.xlsx').active  # loading the excel details file
i=1
if sheet.cell(row=1,column=1).value=='Name':
    i=2
for curr_row in range(i, sheet.max_row + 1):
    img = cv2.imread('template5.jpeg')
    font = cv2.FONT_HERSHEY_DUPLEX  # deciding the font
    text_size = 1  # deciding the font size
    for curr_col in range(1,sheet.max_column+1):
        colu = str(sheet.cell(row=curr_row, column=curr_col).value)
        img = cv2.putText(img, colu, arr[curr_col-1], font, text_size, (0, 0, 200), 1)
    # cv2.imshow('image', img)  # it is showing the certificate
    # can also be commented if you don't wanna see the
    # image windows
    # cv2.waitKey(0)  # making the image window not to disappear until command is given
    dest = dest_path + '/' + str(sheet.cell(row=curr_row,column=1).value) + '.jpeg'  # giving the destination folder where
    # we want to store certificates
    # print(get_name)
    cv2.imwrite(dest, img)  # storing the certificate the destination folder
cv2.destroyAllWindows()

# closing all the windows that opened during the
# execution of this program